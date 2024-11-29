import os
import sys
import PySimpleGUI as sg
import openpyxl as op
#import configparser
import config
from openpyxl.styles import Border, Side, PatternFill, Border

#config = configparser.ConfigParser()
#config.read('config.ini')

sheet_name_exists = True
sample_num_cell = ''
split_x = []

layout = [[sg.Text("Data extraction")],
          [sg.Text("Data files folder"),
           sg.In(size=(25, 1), enable_events=True, key="-FOLDER-"),
           sg.FolderBrowse()],
          [sg.Text("Choose workbook path"),
           sg.In(default_text=config.pathexcel, size=(25, 1), enable_events=True, key="-WB_NAME-"), sg.FilesBrowse(file_types=(("Excel file", "*.xlsx"),))],
          [sg.Text("Enter name of the list"),
           sg.In(size=(25, 1), enable_events=True, key="-LIST_NAME-")],
          [sg.Button('Ok')] ]

window = sg.Window(title='J-V charasteristics extraction program', layout=layout)
while True:
    event, values = window.read()
    if event == 'Ok' or event == sg.WIN_CLOSED:
        break

if values["-FOLDER-"] == 'None' or values["-FOLDER-"] == '':
    window = sg.Window(title='Error', layout=[[sg.Text('Path is not found, try again')], [sg.Button('Ok')]])
    while True:
        event, values = window.read()
        if event == 'Ok' or event == sg.WIN_CLOSED:
            break
    sys.exit()

file_path = str(values["-FOLDER-"])
workbook_path= str(values["-WB_NAME-"])
wb = op.load_workbook(workbook_path)

worksheets = wb.worksheets
sheet_name = str(values["-LIST_NAME-"])
for t in worksheets:
    if t == sheet_name:
        sheet_name_exists = True
        sheet = wb[sheet_name]
        break
    else:
        sheet_name_exists = False

if not sheet_name_exists:
    sheet = wb.create_sheet(sheet_name)

sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

sample_num_cell = sheet.cell(row=1, column=1)
sample_num_cell.value = 'Sample'

voc_cell= sheet.cell(row=1, column=7)
voc_cell.value =  'Voc, V'

isc_cell = sheet.cell(row=1, column=8)
isc_cell.value = 'Isc, mA'

jsc_cell= sheet.cell(row=1, column=9)
jsc_cell.value =  'Jsc, mA/cm2'
jsc_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

ff_cell= sheet.cell(row=1, column=10)
ff_cell.value =  'FF, %'

pce_cell= sheet.cell(row=1, column=11)
pce_cell.value =  'Efficiency, %'

area_cell= sheet.cell(row=1, column=13)
area_cell.value =  'Area, cm2'


row_num = 2
dir_list = os.listdir(file_path)
count = 0

parameters = []
list_len = len(dir_list)

for x in dir_list:
    split_x = x.split('_')
    #print(split_x[0])
    if split_x[0] == 'test':
        #print('ok')
        continue
    elif x == 'test.xlsx':
        #print('this is ' + x)
        continue
    else:
        file = file_path + '\\' + x
        with open(file) as fl:
            lines = fl.read().splitlines()

            voc = [str(r) for r in lines[3].split(',')]
            jsc = [str(y) for y in lines[4].split(',')]
            ff = [str(z) for z in lines[5].split(',')]
            pce = [str(c) for c in lines[6].split(',')]

            parameters = [voc[1].replace(' V =\t', ''), jsc[1].replace(' mA/cm2 =\t', ''),
                    ff[1].replace(' % =\t', ''), pce[1].replace(' % =\t', '')]

            #print(parameters)
            # output:
            # [voc, jsc, ff, pce]


            filename = os.path.basename(file)
            splitted_fln = filename.split(" ")

            sample_perovskite_type = sheet.cell(row=row_num, column=1)
            for unit in splitted_fln:
                for arr in config.perovskite:
                    if unit == arr:
                        sample_perovskite_type.value = unit

            sample_layer_nio_cell = sheet.cell(row=row_num, column=2)
            for x in range(0, len(splitted_fln)):
                if splitted_fln[x] == '2l':
                    sample_layer_nio_cell.value = '2 layers'
                elif splitted_fln[x] == '1l':
                    sample_layer_nio_cell.value = '1 layer'

            sample_num_cell = sheet.cell(row=row_num, column=3)
            for unit in splitted_fln:
                for arr in config.sample_num:
                    if unit == arr:
                        sample_num_cell.value = unit


            pmax_cell = sheet.cell(row=row_num, column=6)
            pmax_cell.value = '=(J' + str(row_num) + '/100)*H' + str(row_num) + '*G' + str(row_num)

            voc_cell= sheet.cell(row=row_num, column=7)
            voc_cell.value =  float(parameters[0])

            isc_cell = sheet.cell(row=row_num, column=8)
            isc_cell.value = '=I' + str(row_num) + '*M' + str(row_num)

            jsc_cell= sheet.cell(row=row_num, column=9)
            jsc_cell.value =  float(parameters[1])
            jsc_cell.fill = PatternFill(start_color="00B0F0",
                                        end_color="00B0F0",
                                        fill_type="solid")

            ff_cell= sheet.cell(row=row_num, column=10)
            ff_cell.value =  float(parameters[2])
            if float(parameters[2]) < 40 or parameters[2] == 'NaN':
                ff_cell.fill = PatternFill(start_color="FF0000",
                                           end_color="FF0000",
                                           fill_type="solid")
            elif float(parameters[2]) > 40 and float(parameters[2]) < 50:
                ff_cell.fill = PatternFill(start_color="FFC000",
                                           end_color="FFC000",
                                           fill_type="solid")
            elif float(parameters[2]) > 50 and float(parameters[2]) < 70:
                ff_cell.fill = PatternFill(start_color="FFFF00",
                                           end_color="FFFF00",
                                           fill_type="solid")

            pce_cell= sheet.cell(row=row_num, column=11)
            pce_cell.value =  float(parameters[3])
            if float(parameters[3]) < 5 or parameters[3] == 'Inf':
                pce_cell.fill = PatternFill(start_color="FF0000",
                                           end_color="FF0000",
                                           fill_type="solid")
            elif float(parameters[3]) > 5 and float(parameters[3]) < 9:
                pce_cell.fill = PatternFill(start_color="FFC000",
                                           end_color="FFC000",
                                           fill_type="solid")
            elif float(parameters[3]) > 9 and float(parameters[3]) < 10:
                pce_cell.fill = PatternFill(start_color="FFFF00",
                                           end_color="FFFF00",
                                           fill_type="solid")
            elif float(parameters[3]) > 10:
                pce_cell.fill = PatternFill(start_color="92D050",
                                           end_color="92D050",
                                           fill_type="solid")

            area_cell = sheet.cell(row=row_num, column=13)
            area_cell.value = 0.21


            row_num += 1


bor_style = Side(border_style='thin', color='000000')
cell_range = sheet['A1':('M' + str(row_num))]
for cell in cell_range:
    for x in cell:
        x.border = Border(top=bor_style,
                          left=bor_style,
                          right=bor_style,
                          bottom=bor_style)

config.change_excel_path(file_path)
wb.move_sheet(sheet_name, offset=-1*len(worksheets))
wb.save(workbook_path)
