import sys
import os
import config
import openpyxl as op
import PySimpleGUI as sg

layout = [[sg.Text("Power and IV data calculation")],
          [sg.Text("Data files folder"),
           sg.In(size=(25, 1), enable_events=True, key="-FOLDER-"),
           sg.FolderBrowse()],
          [sg.Text("Choose workbook path"),
           sg.In(default_text=config.pathexcel, size=(25, 1), enable_events=True, key="-WB_NAME-"), sg.FilesBrowse(file_types=(("Excel file", "*.xlsx"),))],
          [sg.Text("Enter name of the list"),
           sg.In(size=(25, 1), enable_events=True, key="-LIST_NAME-")],
          [sg.Button('Ok')] ]

window = sg.Window(title='Power calculation', layout=layout)
while True:
    event, values = window.read()
    if event == 'Ok' or event == sg.WIN_CLOSED:
        break

if values["-FOLDER-"] == 'None' or values["-FOLDER-"] == '':
    window = sg.Window(title='Error', layout=[[sg.Text('Path is not found, try again')], [sg.Button('Ok')]])
    while True:
        event, values = window.read()
        if event == 'Ok' or event == sg.WIN_CLOSED:
            sys.exit()

file_path = str(values["-FOLDER-"])
workbook_path= str(values["-WB_NAME-"])
wb = op.load_workbook(workbook_path)

worksheets = wb.worksheets
sheet_name = str(values["-LIST_NAME-"])

for t in worksheets:
    if t.title == sheet_name:
        sheet_name_exists = True
        sheet = wb[t.title]
        break
    else:
        sheet_name_exists = False

if not sheet_name_exists:
    sheet = wb.create_sheet(sheet_name)

sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

row_num = 2
count = 0
dir_list = os.listdir(file_path)
list_len = len(dir_list)

for x in dir_list:
    power_list = []
    current_list = []
    voltage_list = []
    i = 3

    split_x = x.split('_')

    if split_x[0] == 'test':
        continue
    elif x == 'test.xlsx':
        continue
    else:
        file = file_path + '\\' + x
        with open(file) as fl:
            lines = fl.read().splitlines()

            for r in lines:
                line = [str(c) for c in lines[i].split('	')]
                voltage = float(line[0])
                current = float(line[1])

                current_list.append(current)
                voltage_list.append(voltage)

                i += 1
                if i == len(lines) - 1:
                    for p in range(0,(len(current_list)-1)):
                        power = current_list[p] * voltage_list[p]
                        power_list.append(power)

                    index = power_list.index(min(power_list))

                    vmax_cell = sheet.cell(row=row_num, column=4)
                    vmax_cell.value = voltage_list[index]

                    imax_cell = sheet.cell(row=row_num, column=5)
                    if current_list[index] > 0:
                        imax_cell.value = current_list[index]
                    elif current_list[index] < 0:
                        imax_cell.value = -1 * current_list[index]

                    r_cell = sheet.cell(row=row_num, column=12)
                    if current_list[index] > 0:
                        r_cell.value = voltage_list[index]/current_list[index]
                    elif current_list[index] < 0:
                        r_cell.value = -1 * voltage_list[index]/current_list[index]

                    list.clear(power_list)
                    list.clear(current_list)
                    list.clear(voltage_list)

                    row_num += 1

                    break

wb.save(workbook_path)
